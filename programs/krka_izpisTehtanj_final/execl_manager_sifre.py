import warnings

from openpyxl import load_workbook

def array_fillS():
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/testing_tehtanje.xlsx", read_only=True)
    ws = wb['vse']
    myArray = []  # B2:B134434
    for row in ws.iter_rows('B2:B1000'):
        # iskani format 03.05.16 11:47:37
        for cell in row:
            myArray.append(cell.value)
    return myArray
