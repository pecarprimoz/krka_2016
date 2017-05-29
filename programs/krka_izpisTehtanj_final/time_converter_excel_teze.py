from collections import defaultdict

from openpyxl import Workbook


# dobim samo dan posebaj
def get_day(time_value):
    parse_day = time_value.split(" ")[0]
    dan = parse_day.split(".")
    return dan[0]


'''
#parser za datum, v tem buildu neuporabljen
def get_date_split(time_value):
    parse_day=time_value.split(" ")[0]
    dan,mesec,leto=int(parse_day.split(".")[0]),int(parse_day.split(".")[1]),int(parse_day.split(".")[2])
    return dan,mesec,leto
'''


# dobim dan + čas
def get_dateT(time_value):
    parse_date = time_value.split(" ")[0:-1]
    return " ".join(parse_date)


# dobim samo čas in ga spremenim v int

def get_time(time_value):
    # break time into 3 parts, take part on ind 1 int(s)
    parse_time = time_value.split(" ")[1]
    h, m, s = parse_time.split(":")
    # change to integers and return
    return int(h), int(m), int(s)


# spremenim vse v sekunde za računanje
def get_seconds(hours, minutes, seconds):
    sumSekunde = 0
    while hours != 0:
        sumSekunde += 3600
        hours -= 1
    while minutes != 0:
        sumSekunde += 60
        minutes -= 1
    return sumSekunde + seconds


# spremenim iz sekund v standarden format
def to_normal(diff):
    ure = 0
    minute = 0
    sekunde = 0
    # vse dobim v sekundah, gledam dokler je večje od 0
    while diff > 0:
        if minute >= 60:
            minute -= 60
            ure += 1
        if sekunde >= 60:
            sekunde -= 60
            minute += 1
        if diff % 3600 == 0:
            ure += 1
            diff -= 3600
        if diff % 60 == 0:
            minute += 1
            diff -= 60
        else:
            diff -= 1
            sekunde += 1
    return "{:02}:{:02}:{:02}".format(ure, minute, sekunde)


# =""&B1&","
# lifesaver na začetku, sedaj zamenjal na openpyxl
def core(sezCasov, sezCifer):
    # naredim nov virutalen workbook, v njemu sheet "Podatki med izmeno"
    wb = Workbook()
    ws = wb.active
    ws.title = "Podatki med izmeno"
    # counter za celice
    counter = 2
    # dict v katerem imam datume:čase
    myDt = defaultdict(list)
    for ind, cas in enumerate(sezCasov):
        # preverimo da ne dobimo index out of range errora
        if ind != len(sezCasov) - 1:
            # poberemo oba dneva za primerjanje ali smo že na novem dnevu
            trenutenDan = get_day(cas)
            naslednjiDan = get_day(sezCasov[ind + 1])

            # poberem dan in čas za v dict
            trenutenDatum, trenutenCas = get_dateT(cas).split(" ")[0], get_dateT(cas).split(" ")[1]

            # če je šifra za tehtanje pravilna dodamo v dict, s tem potrdimo tehtanje
            if sezCifer[ind] == 5110:
                # če še ni v dictu ga dodamo z enim časom
                if trenutenDatum not in myDt:
                    myDt[trenutenDatum].append(trenutenCas)
            if trenutenDatum in myDt:
                # v primeru da je v dictu in da se dan razlikuje dodamo še končni čas
                if trenutenDan != naslednjiDan:
                    myDt[trenutenDatum].append(trenutenCas)
        # exception za zadnjega
        if ind == len(sezCasov) - 1:

            myDt[trenutenDatum].append(trenutenCas)
    # sprememba string datuma v datetime, hotel bi imeti boljšo rešitev, saj sedaj sortiram v excelu
    import datetime
    for key, value in myDt.items():
        # shranimo si samo datum
        curDate = datetime.datetime.strptime(key, "%d.%m.%Y").date()
        # formatiranje, v primeru da nimamo podanega časa je problem v podatkih, dodamo samo tiste,
        # kjer imamo začetni in končni čas
        if len(value) == 2:
            cellB = "Prva meritev: {}      Zadnja meritev: {}".format(value[0], value[1])
            ws['B' + str(counter)] = cellB
        ws['A' + str(counter)] = curDate
        ws['C' + str(counter)] = value[0]  # prva meritev
        # samo minute in sekunde
        minSec = "00:{}:{}".format(value[0].split(":")[1], value[0].split(":")[2])
        ws['D' + str(counter)] = minSec  # minute in sekunde
        ws['E' + str(counter)] = value[1]  # zadnja meritev
        # odmik od naslednje polne ure
        # vzamem za prvega eno uro za drugega pa samo minute in sekunde brez ur
        sekundaUre = get_seconds(1, 0, 0)
        sekundaOstalo = get_seconds(0, int(value[0].split(":")[1]), int(value[0].split(":")[2]))
        curRazlika = sekundaUre - sekundaOstalo
        ws['F' + str(counter)] = to_normal(curRazlika)
        # povečamo counter da se premikamo po željenih celicah
        counter += 1
    # shranimo datoteko

    wb.save("Prvo in zadnje tehtanje.xlsx")


# zaženemo funkcijo ki poganja vse drugo v initu
from krka_izpisTehtanj_final.time_converter_excel_teze import core
from krka_izpisTehtanj_final.excel_manager_tehtanje import array_fillT
from krka_izpisTehtanj_final.execl_manager_sifre import array_fillS

core(array_fillT(), array_fillS())
