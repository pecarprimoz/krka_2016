import warnings

from openpyxl import load_workbook


# izdelava arraja datumov za tehtanje
def array_fillT():
    warnings.simplefilter("ignore")
    # uporabimo excel datoteko testing_tehtanje, read only za dodatno hitrost
    wb = load_workbook(filename="C:/Podatki/testing_tehtanje.xlsx", read_only=True)

    # sheet z imenom 'vse'
    ws = wb['vse']
    myArray = []
    for row in ws.iter_rows('A2:A1000'):
        # iskani format 03.05.16 11:47:37
        # datetime omogoča klicanje vsakega dela datuma posebaj
        for cell in row:
            myDt = cell.value
            leto = myDt.year
            mesec = myDt.month
            dan = myDt.day
            h = myDt.hour
            m = myDt.minute
            s = myDt.second
            dateFormatCorrect = "{D:02}.{Me:02}.{L} {H:02}:{Mi:02}:{S:02} TEMP".format(L=leto, Me=mesec, D=dan,
                                                                                       H=h, Mi=m, S=s)
            # formatiram datum za uporabo v glavnem programu in ga dam v array
            myArray.append(dateFormatCorrect)
    return myArray
