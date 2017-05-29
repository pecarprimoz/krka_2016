import warnings

from openpyxl import Workbook, load_workbook


# TODO boš še jutr baje rabu če se ne compila si neki zajebu k si spreminju na drugem file-u
# kjer je bil problem z formulami (znebu z raw data, pr taprvm je magično delal?)
################# BANGER KOMANDA################
#### =IF(A2="NEKAJ";"sTRING";A2/B2)         ####
#### IF COLUMN EQUALS SOMETHING; TRUE; FALSE####
################################################

def getdata():
    # A2-A255736
    warnings.simplefilter("ignore")  # raw cut pa data only te reš pred retardiranimi funkcijami
    # thank fuck za komentarje
    wb = load_workbook(filename="C:/Podatki/le_razlika1.xlsx", read_only=True, data_only=True)
    ws = wb['rzl']
    razl = []
    undef = []
    # vse podatke k so že dokončni dam v tabelo razl
    for row in ws.iter_rows('C2:C3819'):
        for cell in row:
            razl.append(cell.value)
    # vsi podatki ki potrebujejo obdelavo damo v undef
    for row in ws.iter_rows('A2:A3819'):
        for cell in row:
            undef.append(cell.value)
    # vrnemo oboje
    return undef, razl


def turnup_dif(narobe):
    # tiste ki so narobni jih spremenimo
    for ind, el in enumerate(narobe):
        # če ni string gremo naprej
        if not isinstance(el, str):
            # spremenimo v string in probamo najti piko, če je ne najdemo pomen da je celo število
            tmp = str(el)
            to = str(tmp).find(".")
            if to != -1:
                # torej če je decimalna vejica vzamemo celoto in prištejemo eno, to zapišemo na isti indeks
                pristeto = int(tmp[:to]) + 1
                narobe[ind] = pristeto
            else:
                # v primeru da je celo dodamo 1, idk če je pravilno #todo lol
                pristeto = int(tmp) + 1
                narobe[ind] = pristeto
    return narobe


def odstej(teo, pra):
    wb = Workbook()
    ws = wb.active
    ws.title = "razlika"
    # oba imam preurejena, grem čez tistega ki je vredu in zapisujemo razliko v posebaj file
    for ind, val in enumerate(teo):
        # ker smo prej not pustil vrednosti (da ohranimo pravilni vrstni red)
        # samo preverimo da stvar ni string (torej neki tacga. #NULLFCK) in zapišemo na določeno mesto
        if not isinstance(val, str):
            # ind +2 zaradi tega ker začnemo na 0 in ker je 1 column ime stolpca
            ws['A' + str(ind + 2)] = val - pra[ind]
    # shranimo v poljubno datoteko
    wb.save("RazlikaNew.xlsx")


# shranimo v posebaj datoteke
a, b = getdata()
# obdelamo float valu-e in jih shranimo nazaj
sp = turnup_dif(a)
# odštejemo in shranimo
odstej(sp, b)
