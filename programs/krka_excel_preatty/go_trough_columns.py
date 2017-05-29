import warnings
from collections import *
from random import randint

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def random_color():
    arrayChr = []
    for i in range(7):
        arrayChr.append(chr(randint(65, 70)))
    intChr = []
    for i in range(7):
        intChr.append(i)

    finalStr = ""

    for i in range(6):
        mem = randint(1, 2)
        rnd = randint(0, 6)
        if mem == 1:
            finalStr += str(arrayChr[rnd])
        else:
            finalStr += str(intChr[rnd])
    print(finalStr)
    return finalStr


def array_values():
    blueFill = PatternFill(start_color='B3E5FC',
                           end_color='B3E5FC',
                           fill_type='solid')
    greenFill = PatternFill(start_color='FBE9E7',
                            end_color='FBE9E7',
                            fill_type='solid')

    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/herb_koncano_testi.xlsx")
    ws = wb['Podatki']
    myDict = defaultdict(list)
    for ind, row in enumerate(ws.iter_rows('A2:A537')):
        for cell in row:
            if cell.value == "1000" or cell.value == 1000:
                for i in range(65, 91):
                    ws[chr(i) + str(ind + 2)].fill = PatternFill(start_color=random_color(), fill_type="solid")
            elif cell.value is None:
                for i in range(65, 91):
                    ws[chr(i) + str(ind + 2)].fill = PatternFill(start_color=random_color(), fill_type="solid")

    wb.save("testing.xlsx")


print(array_values())
