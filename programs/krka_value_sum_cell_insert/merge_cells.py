import warnings

from openpyxl import Workbook, load_workbook


# napisano samo da da skupaj vrednosti celic, ker je excel posebna princesa in tega ni zmožen
def getCode():
    # A2-A255736
    warnings.simplefilter("ignore")  # lahko samo fileformat spreminjaš v primeru da je postavitev ista
    wb = load_workbook(filename="C:/Podatki/tretji.xlsx", read_only=True)
    ws = wb['tr']
    codeArray = []
    otherArray = []
    # A2:A537#A2:A255736
    for row in ws.iter_rows('A2:A255736'):
        for cell in row:
            codeArray.append(cell.value)
    for row in ws.iter_rows('B2:B255736'):
        for cell1 in row:
            otherArray.append(cell1.value)
    # tam kjer je vrednost None, torej prazna celica; zamenjamo z vrednostjo v drugem arrayu(torej ostale vrednosti)
    for ind, val in enumerate(codeArray):
        if val == None:
            codeArray[ind] = otherArray[ind]
    return codeArray


def spremeni(cod):
    wb = Workbook()
    ws = wb.active
    ws.title = "pls work"
    for ind, val in enumerate(cod):
        # dodamo vrednost na določeno mesto, DELAŠ 2 PA PO 2, TOREJ ŠIFRA=>TEH1, ŠIFRA=>TEH2 ect.
        ws['A' + str(ind + 2)] = val

    wb.save("Tretji.xlsx")


spremeni(getCode())
