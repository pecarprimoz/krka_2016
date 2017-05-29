import warnings
from collections import *

from openpyxl import Workbook, load_workbook


def getCode():
    # A2-A255736
    # ignoriram warninge poberem use šifre in jih vrnem
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/samSifre.xlsx", read_only=True)
    ws = wb['sifre']
    codeArray = []                        # A2:A537#A2:A255736
    for ind, row in enumerate(ws.iter_rows('A2:A255736')):
        for cell in row:
            codeArray.append(cell.value)
    return codeArray


def getOne():
    # A2-A255736
    # poberm vse vrednosti od prve tehtalnice
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/one.xlsx", read_only=True)
    ws = wb['one']
    codeArray = []  # B2:B537#B2:B255736
    for ind, row in enumerate(ws.iter_rows('A2:A255736')):
        for cell in row:
            codeArray.append(cell.value)
    return codeArray


def getThree():
    # poberm vse vrednosti od tretje tehtalnice
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/three.xlsx", read_only=True)
    ws = wb['three']
    codeArray = []  # C2:C537#C2:C255736
    for row in ws.iter_rows('A2:A255736'):
        for cell in row:
            codeArray.append(cell.value)
    return codeArray


def getFive():
    # poberm vse vrednosti od pete tehtalnice
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/five.xlsx", read_only=True)
    ws = wb['five']
    codeArray = []  # D2:D537#D2:D255736
    for row in ws.iter_rows('A2:A255736'):
        for cell in row:
            codeArray.append(cell.value)
    return codeArray


def sumOne(codeArray, oneArray, threeArray, fiveArray):
    # po šifrah seštejem za vsako tehtalnico vrednosti, shranim v dict
    myDtOne = defaultdict(float)
    myDtThree = defaultdict(float)
    myDtFive = defaultdict(float)
    for valOne in zip(codeArray, oneArray):
        myDtOne[valOne[0]] += float(valOne[1])
    for valThree in zip(codeArray, threeArray):
        myDtThree[valThree[0]] += float(valThree[1])
    for valFive in zip(codeArray, fiveArray):
        myDtFive[valFive[0]] += float(valFive[1])
    finalDict = defaultdict(list)
    # vse dam v en dict v formatu šifra:[teh1,teh3,teh5]
    for key, valOne in myDtOne.items():
        # key je šifra, val je seštevek
        finalDict[key].append(valOne)
    for key, valThree in myDtThree.items():
        # key je šifra, val je seštevek
        finalDict[key].append(valThree)
    for key, valFive in myDtFive.items():
        # key je šifra, val je seštevek
        finalDict[key].append(valFive)

    return finalDict


def replaceAllOccurances(sifreMainFile, arrayDicVse):
    wb = Workbook()
    ws = wb.active  # TODO lol
    ws.title = "pls work"
    # grem cez vse šifre
    for key, value in arrayDicVse.items():
        # grem samo čez šifre in iščem zadetek, s tem dobim ideks kamor bom pisal
        for ind, i in enumerate(sifreMainFile):
            if i == key:
                '''
                imam dict ki ne drži vrstnega reda, ta ima seštete vrednosti,
                ko grem čez main file oz. čez sifre main file-a mu moram dodat vrednosti
                na točno določen indeks, spustim 0 in 1 torej vedno gledam ind+2

                '''
                # vse tri vrednosti vnesem zgoraj
                one, three, five = value[0], value[1], value[2]
                ws['C' + str(ind + 2)] = one
                ws['D' + str(ind + 2)] = three
                ws['E' + str(ind + 2)] = five
                # print("ok")
                break
    # vrnem in potem sestavim
    wb.save("Zelo veliko podatkov.xlsx")


onlyCode = getCode()
# allValues=sumOne(onlyCode,getOne(),getThree(),getFive())
replaceAllOccurances(onlyCode, sumOne(onlyCode, getOne(), getThree(), getFive()))
