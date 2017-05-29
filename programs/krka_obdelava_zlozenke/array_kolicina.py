import warnings

from openpyxl import load_workbook


# dobim količino iz columna b


def array_kolicina():
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/newZlozenke.xlsx", read_only=True)
    ws = wb['Podatki']
    myArray = []
    for row in ws.iter_rows('C2:C20477'):
        for cell in row:
            myArray.append(cell.value)
    return myArray
