from openpyxl import Workbook

from krka_obdelava_zlozenke.array_kolicina import array_kolicina
from krka_obdelava_zlozenke.excel_material_getter import array_materials
from krka_obdelava_zlozenke.excel_value_getter import array_values


def core(material_array, value_array, kolicina_array):
    # naredim nov virutalen workbook, v njemu sheet "Podatki med izmeno"
    wb = Workbook()
    ws = wb.active
    ws.title = "Podatki za tehtanje"
    # counter za celice
    vsota = 0
    countSifre = 0
    kolicinaSum = 0
    for ind, skupno in enumerate(zip(material_array, value_array)):
        # materialSIF = šifra materiala ki ga računamo
        # valueSIF = dejanski znesek katerega bomo seštevali
        materialSIF, valueSIF = skupno
        if ind != len(material_array) - 1:
            if materialSIF != material_array[ind + 1]:
                vsota += float(valueSIF)
                countSifre += 1
                kolicinaSum += kolicina_array[ind]
                # če se zamena mora indeks vedno biti 2 večji 0- ni v exc, 1- je vendar je neki X
                # vsota zneskov
                ws['A' + str(ind + 2)] = vsota
                # ponovitev sifer
                ws['B' + str(ind + 2)] = countSifre
                # skupna kolicina
                ws['C' + str(ind + 2)] = kolicinaSum
                # ker je zadnja vrednost postavim na 0
                vsota = 0
                countSifre = 0
                kolicinaSum = 0
            # če ni zadnja vrednost samo povečam spremenljivke
            elif materialSIF == material_array[ind + 1]:
                vsota += float(valueSIF)
                countSifre += 1
                kolicinaSum += kolicina_array[ind]
        # ko smo na zadnem indeksu dodamo zapis in končamo
        else:
            vsota += float(valueSIF)
            countSifre += 1
            kolicinaSum += kolicina_array[ind]
            # če se zamena mora indeks vedno biti 2 večji 0- ni v exc, 1- je vendar je neki X
            ws['A' + str(ind + 2)] = vsota
            ws['B' + str(ind + 2)] = countSifre
            ws['C' + str(ind + 2)] = kolicinaSum
            vsota = 0
            countSifre = 0
            kolicinaSum = 0

    wb.save("FirstRun.xlsx")


core(array_materials(), array_values(), array_kolicina())
