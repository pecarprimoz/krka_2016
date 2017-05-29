import warnings

from openpyxl import load_workbook


# dobim material iz columna A
def array_materials():
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/newZlozenke.xlsx", read_only=True)
    ws = wb['Podatki']
    myArray = []
    for row in ws.iter_rows('A2:A20477'):
        for cell in row:
            myArray.append(cell.value)
    return myArray
