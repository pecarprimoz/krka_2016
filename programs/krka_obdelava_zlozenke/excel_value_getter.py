import warnings

from openpyxl import load_workbook


# dobim denar
def array_values():
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/newZlozenke.xlsx", read_only=True)
    ws = wb['Podatki']
    myArray = []
    for row in ws.iter_rows('B2:B20477'):
        for cell in row:
            myArray.append(cell.value)
    return myArray
