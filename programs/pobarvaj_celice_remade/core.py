import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

def poberi_podatke():
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/procenti2.xlsx", read_only=True)
    ws = wb['mydata']
    #65 je a, prevert morm kje se končajo vnosi
    koncajo=0
    #ASCII reprezentacija stevil od A do Z po anglesko
    for i in range(65,91):
        if ws[chr(i)+"1"].value is None:
            #dokler ne pride do value-a ki je none
            koncajo=i
            break
    arVseh=[[] for j in range(koncajo-65)]
    counter=0
    for k in range(65,koncajo):
        #dodajamo v arraye ki so toliko dolgi kot koncajo var
        for ind, row in enumerate(ws.iter_rows(chr(k)+'1:'+chr(k)+'18')):
            #dodajamo v array
            for cell in row:
                arVseh[counter].append(cell.value)
        counter+=1

    return arVseh

def rework_data(array):
    for skup in array:
        for ind,value in enumerate(skup):
            if value is None:
                #ce je cell value bil None v array damo 0 za kasnejso racunanje
                skup[ind] = 0
            if isinstance(value, str):
                #pogledamo kje je vejica ker je tako shranjeno v excelu, da mejamo z .
                kje = value.find(",")
                if kje != -1:
                    mojList = list(value)
                    mojList[kje] = "."
                    skup[ind] = "".join(mojList)
    return array
def remove_strings(array):
    temp=array.copy()
    katereVn=[]
    for ind,poseb in enumerate(temp):
        #ce ne najdemo pike dodamo indeks v array
        if poseb[0].find(".")==-1:
            katereVn.append(ind)
    for i in reversed(katereVn):
        #na podlagi arraya grajenega zgoraj popamo od zadaj ven, no erors WEW
        array.pop(i)
    return array
def predelaj_podatke(array):
    wb = Workbook()
    ws = wb.active
    zacetek=65
    ws.title = "ProcenteRihtam"
    counter = 1
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    for skup in array:
        #average in counter za average
        avg=0
        avc=0
        #spremenimo prvo celico v celoto
        celota = float(skup[0])
        procentualno = celota * 0.3
        #dobimo max in min iz celote, jemljemo po 30% gor/dol
        max = celota + procentualno
        min = celota - procentualno
        for ind, vrednost in enumerate(skup[1:]):
            #spremenimo v float, dodamo v average, se je v polju vecjem ali manjsem
            #obarvamo celico na rdece
            vrednost1 = float(vrednost)
            print(vrednost1)
            avg+=vrednost1
            if vrednost1 > max or vrednost1 < min:
                ws[chr(zacetek) +  str(counter)] = vrednost1
                ws[chr(zacetek) +  str(counter)].fill = redFill
            else:
                ws[chr(zacetek) +  str(counter)] = vrednost1
            counter += 1
            avc+=1
        #na koncu vsakega zapišemo povprečje vseh celic za to vrstico, in postavimo vse na 1

        ws[chr(zacetek) + str(counter+1)] = avg/(avc)
        counter=1
        zacetek+=1
        print(counter)
        print(zacetek)
    wb.save("CUSTOM PROCENTI.xlsx")

fillArray=poberi_podatke()
revArray=rework_data(fillArray)
koncanArray=remove_strings(revArray)
predelaj_podatke(koncanArray)