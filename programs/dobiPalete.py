import warnings
from openpyxl import Workbook, load_workbook

def getPodatke():
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/plsf.xlsx", read_only=True)
    ws = wb['mydata']
    cola=[]
    colb=[]
    for row in ws.iter_rows('A2:A15712'):
        # dodajamo v array
        for cell in row:
            cola.append(cell.value)
    for row1 in ws.iter_rows('B2:B15713'):
        # dodajamo v array
        for cell1 in row1:
            colb.append(cell1.value)

    return cola,colb

from collections import defaultdict
def predelaj(a,b):
    myDt=defaultdict(int)
    for skupno in zip(a,b):
        sifra,palete=skupno
        if palete is None:
            palete=0
        myDt[sifra]+=int(palete)
    return myDt

def core(dct):
    wb = Workbook()
    ws = wb.active
    counter=1

    ws.title = "ProcenteRihtam"
    for key,value in dct.items():
        ws['A'+str(counter)] = key
        ws['B' + str(counter)] = value
        counter+=1

    wb.save("firstTry.xlsx")
a,b=getPodatke()
myDt=predelaj(a,b)
core(myDt)
