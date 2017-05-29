import warnings

from openpyxl import load_workbook


# first revision, time format ex. 03.05.16 11:47:37 EUROPE/LJUBLJANA
def array_fillD():
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/newFiles_tehtanje.xlsx", read_only=True)
    ws = wb['Podatki']
    myArray = []
    for row in ws.iter_rows('A63:A10000'):
        for cell in row:
            myArray.append(cell.value)
    return myArray


# second revision, 130k cols, time format ex. 9.7.2016  12:11:37 - datetime(2016,7,9,12,11,37)
'''

def array_fill():
    warnings.simplefilter("ignore")
    wb=load_workbook(filename="C:/Podatki/myDataB.xlsx",read_only=True)
    ws=wb['vse']
    myArray=[]
    for row in ws.iter_rows('A2:A134434'):
        #iskani format 03.05.16 11:47:37
        for cell in row:
            myDt=cell.value
            leto=myDt.year
            mesec=myDt.month
            dan=myDt.day
            h=myDt.hour
            m=myDt.minute
            s=myDt.second
            dateFormatCorrect="{D:02}.{Me:02}.{L} {H:02}:{Mi:02}:{S:02} TEMP".format(L=leto,Me=mesec,D=dan,
                                                                      H=h,Mi=m,S=s)
            myArray.append(dateFormatCorrect)
    return myArray

array_fill()
'''
