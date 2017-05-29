# excel writer za gledanje izmen med 13 in 14, druga izmena
from openpyxl import Workbook
from openpyxl.styles import PatternFill


# dobim samo dan posebaj
def get_day(time_value):
    parse_day = time_value.split(" ")[0]
    dan = parse_day.split(".")
    return dan[0]


# dobim dan + čas
def get_dateT(time_value):
    parse_date = time_value.split(" ")[0:-1]
    return " ".join(parse_date)


# dobim samo čas in ga spremenim v int
def get_time(time_value):
    # break time into 3 parts, take part on ind 1 int(s)
    parse_time = time_value.split(" ")[1]
    h, m, s = parse_time.split(":")
    # change to integers and return
    return int(h), int(m), int(s)


# spremenim vse v sekunde za računanje
def get_seconds(hours, minutes, seconds):
    sumSekunde = 0
    while hours != 0:
        sumSekunde += 3600
        hours -= 1
    while minutes != 0:
        sumSekunde += 60
        minutes -= 1
    return sumSekunde + seconds


# spremenim iz sekund v standarden format
def to_normal(diff):
    ure = 0
    minute = 0
    sekunde = 0
    while diff > 0:
        if minute >= 60:
            minute -= 60
            ure += 1
        if sekunde >= 60:
            sekunde -= 60
            minute += 1
        if diff % 3600 == 0:
            ure += 1
            diff -= 3600
        if diff % 60 == 0:
            minute += 1
            diff -= 60
        else:
            diff -= 1
            sekunde += 1
    return "{:02}:{:02}:{:02}".format(ure, minute, sekunde)


########################################
##ZELO UPORABA KOMANDA DONT LOOSE THIS##
##="'"&B1&"',"                        ##
########################################
# glavni del programa, podamo mu array stringov(tukaj bom moral zamenjati z columni ko bo dejanska uporaba v excelu)


def core(sezCasov, sezSource):
    # naredim workbook v katerega bom pisal
    wb = Workbook()
    ws = wb.active
    ws.title = "Podatki med izmeno"
    counter = 0
    # naredim si pattern barve katerega bom uporabil pod določenem pogoju
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    # grem čez array datumov in časa
    for ind, bigVal in enumerate(sezCasov):
        # preverimo da ni index out of bound errora
        import datetime
        if ind != len(sezCasov) - 1:
            dateFirst = get_dateT(bigVal)
            # splittat in vzeti prvega
            curDate = str(dateFirst.split(" ")[0])
            # dobim date za prvi stolpec
            objDate = datetime.datetime.strptime(curDate, "%d.%m.%Y").date()
            # dobim celoten datum za izpis(grde stvari so tut zravn se da popravit)
            dateSecond = get_dateT(sezCasov[ind + 1])
            # dobim ure minute in sekunde posebaj za oba časa
            h1, m1, s1 = get_time(bigVal)
            h2, m2, s2 = get_time(sezCasov[ind + 1])
            # poberem samo prvi dan da preverjam če je nov dan
            danPrvi = get_day(bigVal)
            danDrugi = get_day(sezCasov[ind + 1])
            # dobim sekunde za razliko
            casPrvi = get_seconds(h1, m1, s1)
            casDrugi = get_seconds(h2, m2, s2)
            razlika = to_normal(casDrugi - casPrvi)
        # če je dan drugega večji kot dan prvega pomeni da je sedaj drugačen datum
        if danDrugi > danPrvi:
            print("-------------------------- NOV DATUM -------------------------\n")
        else:
            # če je čas ena 14, in čas dva karkoli večjega (menjava izmene) pomeni da računamo razliko
            if h1 == 13 and h2 >= 14 or h1 == 12 and h2 >= 14:  # spremenjeno na vse kar je po 14 se sprejme kot pravilno
                counter += 1
                cellB = "Razlika med {F} in {S}".format(F=dateFirst, S=dateSecond)
                cellC = razlika
                # dodam not datetime(datum)
                ws['A' + str(counter)] = objDate
                # dodam notr raw zapis
                ws['B' + str(counter)] = cellB
                # dodam not razliko
                ws['C' + str(counter)] = cellC
                # dodam not uro prvega tehtanja
                ws['D' + str(counter)] = dateFirst.split(" ")[1]
                # dodam not uro drugega tehtanja
                ws['E' + str(counter)] = dateSecond.split(" ")[1]
                # dodam not šifro tehtalnice
                ws['F' + str(counter)] = sezSource[ind]
                # preverim če je razlika minut nad 45 ali 1 uro; označim z rdečo če je
                if int(razlika.split(":")[1]) > 45 or int(razlika.split(":")[0]) >= 1:
                    ws['C' + str(counter)].fill = redFill
                print("Razlika med {F} in {S} je {R}\n".format(F=dateFirst, S=dateSecond, R=razlika))
    wb.save("Podatki med prvo in drugo izmeno_verzija2.xlsx")


# zaženemo funkcijo ki poganja vse drugo v init-u
from krka_izpisProizvodne.time_converter_excel_writer_izmene import core
from krka_izpisProizvodne.excel_date_getter import array_fillD
from krka_izpisProizvodne.excel_source_getter import array_fillS

core(array_fillD(), array_fillS())
