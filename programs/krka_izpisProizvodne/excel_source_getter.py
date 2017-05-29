import warnings

from openpyxl import load_workbook


def array_fillS():
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/newFiles_tehtanje.xlsx", read_only=True)
    ws = wb['Podatki']
    myArray = []
    for row in ws.iter_rows('B63:B10000'):
        for cell in row:
            myArray.append(cell.value)
    return myArray
