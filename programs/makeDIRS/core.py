import warnings
from openpyxl import load_workbook
import os
def getPodatke():
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/vsi INN + prevladujoča blagovna znamka.xlsx", read_only=True)
    ws = wb['mydt']
    cola=[]
    colb=[]
    for row in ws.iter_rows('A2:A294'):
        for cell in row:
            cola.append(cell.value)
    for row1 in ws.iter_rows('B2:B294'):
        for cell1 in row1:
            colb.append(cell1.value)
    return cola,colb
def vse(a,b):
    for skupno in zip(b,a):
        prvi=skupno[0]
        drugi=skupno[1]
        endStr=drugi+"_"+prvi
        if not os.path.exists("C:/Podatki/koncano/"+endStr):
            os.makedirs("C:/Podatki/koncano/"+endStr)
a,b=getPodatke()
vse(a,b)