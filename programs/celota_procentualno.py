import warnings
from collections import *

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

def getCode():
    # A2-A255736
    # ignoriram warninge poberem use šifre in jih vrnem
    warnings.simplefilter("ignore")
    wb = load_workbook(filename="C:/Podatki/procenti.xlsx", read_only=True)
    ws = wb['mydata']
    A = []                        # A2:A537#A2:A255736
    B = []
    C = []
    D = []
    E = []
    F = []
    G = []
    H= []
    I= []
    J= []
    K= []
    L= []
    M= []
    N= []
    O= []
    P= []
    Q= []
    R= []
    S= []
    T= []


    for ind, row in enumerate(ws.iter_rows('A1:A43')):
        for cell in row:
            A.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('B1:B43')):
        for cell in row:
            B.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('C1:C43')):
        for cell in row:
            C.append(cell.value)

    for ind, row in enumerate(ws.iter_rows('D1:D43')):
        for cell in row:
            D.append(cell.value)

    for ind, row in enumerate(ws.iter_rows('E1:E43')):
        for cell in row:
            E.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('F1:F43')):
        for cell in row:
            F.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('G1:G43')):
        for cell in row:
            G.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('H1:H43')):
        for cell in row:
            H.append(cell.value)

    for ind, row in enumerate(ws.iter_rows('I1:I43')):
        for cell in row:
            I.append(cell.value)

    for ind, row in enumerate(ws.iter_rows('J1:J43')):
        for cell in row:
            J.append(cell.value)

    for ind, row in enumerate(ws.iter_rows('K1:K43')):
        for cell in row:
            K.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('L1:L43')):
        for cell in row:
            L.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('M1:M43')):
        for cell in row:
            M.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('N1:N43')):
        for cell in row:
            N.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('O1:O43')):
        for cell in row:
            O.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('P1:P43')):
        for cell in row:
            P.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('Q1:Q43')):
        for cell in row:
            Q.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('R1:R43')):
        for cell in row:
            R.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('S1:S43')):
        for cell in row:
            S.append(cell.value)
    for ind, row in enumerate(ws.iter_rows('T1:T43')):
        for cell in row:
            T.append(cell.value)
    return A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T


def zamenajVejce(array):
    for ind,value in enumerate(array):
        if value is None:
            array[ind]=0
        if isinstance(value,str):
            temp=value
            kje=value.find(",")
            if kje!=-1:
                mojList=list(value)
                mojList[kje]="."
                array[ind]="".join(mojList)
    return array

zacetek=65
def getPrecentages(array,num):
    wb = Workbook()
    ws = wb.active
    ws.title = "ProcenteRihtam"
    counter = 1
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    celota=float(array[0])
    procentualno=celota*0.3
    max=celota+procentualno
    min=celota-procentualno
    for ind, vrednost in enumerate(array[1:]):
        vrednost1=float(vrednost)
        if vrednost1>max or vrednost1<min:
            ws['A' + str(counter)] = vrednost1
            ws['A' + str(counter)].fill = redFill
        else:
            ws['A' + str(counter)] = vrednost1
        counter+=1
    wb.save(str(num)+"CUSTOM PROCENTI.xlsx")
    #, E, F, G, C, D
A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T=getCode()
sA=zamenajVejce(A)
sB=zamenajVejce(B)
sC=zamenajVejce(C)
sD=zamenajVejce(D)
sE=zamenajVejce(E)
sF=zamenajVejce(F)
sG=zamenajVejce(G)
sH=zamenajVejce(H)
sI=zamenajVejce(I)
sJ=zamenajVejce(J)
sK=zamenajVejce(K)
sL=zamenajVejce(L)
sM=zamenajVejce(M)
sN=zamenajVejce(N)
sO=zamenajVejce(O)
sP=zamenajVejce(P)
sQ=zamenajVejce(Q)
sR=zamenajVejce(R)
sS=zamenajVejce(S)
sT=zamenajVejce(T)

getPrecentages(sA,1)
getPrecentages(sB,2)
getPrecentages(sC,3)
getPrecentages(sD,4)
getPrecentages(sE,5)
getPrecentages(sF,6)
getPrecentages(sG,7)
getPrecentages(sH,8)
getPrecentages(sI,9)
getPrecentages(sJ,10)
getPrecentages(sK,11)
getPrecentages(sL,12)
getPrecentages(sM,13)
getPrecentages(sN,14)
getPrecentages(sO,15)
getPrecentages(sP,16)
getPrecentages(sQ,17)
getPrecentages(sR,18)
getPrecentages(sS,19)
getPrecentages(sT,20)
#fukin lmao what were you thinking
